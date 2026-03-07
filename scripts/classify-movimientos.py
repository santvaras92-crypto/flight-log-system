#!/usr/bin/env python3
"""
Classify all rows in Movimientos.xlsx with one of 12 tipo categories.

Categories:
1.  Pago piloto     – pilot deposits (ingresos with pilot code)
2.  Combustible     – fuel purchases
3.  Mantenimiento   – service/labor (Rev 100, cambio aceite, lavado, mecánico)
4.  Repuestos       – parts/equipment purchases
5.  Hangar          – monthly hangar rent
6.  Seguro          – insurance
7.  Overhaul        – engine/propeller overhaul
8.  Inversión       – Fintual, DAP, Fondo Mutuo, Deposito a plazo
9.  Impuesto        – TOA, Patente Municipal, impuestos
10. Banco           – Comisiones, IVA Comisiones
11. Operacional     – Contador, SyP, misc operating expenses
12. Sin clasificar  – Ceres, Aerotrust, anything unclear
"""

import openpyxl
import re
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'Cuenta banco', 'Movimientos.xlsx')

# Known pilot codes (cliente column)
PILOT_CODES = {
    'SV', 'JP', 'AC', 'FP', 'JR', 'IR', 'RL', 'CP', 'AGH', 'IO',
    'MC', 'ML', 'RM', 'JPB', 'DB', 'VB', 'EA', 'PV', 'GC', 'VBA',
    'IC', 'CR', 'NR', 'RB', 'NES', 'MLU', 'FC', 'JA', 'GT', 'MCA',
    'IA', 'MA', 'BJ', 'SM', 'CF', 'CV', 'CVA', 'DY', 'RG', 'MH',
    'GG', 'ES', 'SE', 'JL', 'MG', 'VO', 'TG', 'SA', 'FL', 'JPL',
    'PG', 'MM', 'SN', 'CM', 'DC', 'MO', 'FM', 'MV', 'TGCO', 'CRO',
    'AP', 'PS', 'NL', 'MOS', 'SC', 'FE', 'JCD', 'MS', 'ED', 'BR',
    'DV', 'ICI', 'EAL', 'NP', 'JO', 'MD', 'SO', 'FG', 'AM', 'PSA',
    'LM', 'DL', 'FA', 'MV', 'PJ', 'AS', 'DO', 'DRCO', 'PA',
}

# Two-letter pilot codes that appear in cliente column
# Some entries have Fpro, USD, Cer, Stratus — those are NOT pilot codes

def classify_row(desc: str, egreso, ingreso, tipo_existing: str, cliente: str) -> str:
    """Return the tipo classification for a row."""
    desc_lower = desc.lower().strip() if desc else ''
    cliente_clean = (cliente or '').strip()

    # ── 0. Already has a correct tipo? Only keep Combustible/Mantenimiento that were manually set ──
    # We reclassify everything to apply updated rules
    # (The valid_tipos check is disabled for a full reclassification pass)

    # ── 1. BANCO ──
    if re.search(r'^comisiones$', desc_lower):
        return 'Banco'
    if re.search(r'^iva comisiones$', desc_lower):
        return 'Banco'

    # ── 2. HANGAR ──
    if re.search(r'hangar|pago estacionamiento', desc_lower):
        # "Trabajos Hangar Gabriel" → Operacional, not Hangar
        if re.search(r'trabajos.*hangar|hangar.*trabajos|gabriel.*hangar|limpieza hangar|movimiento.*hangar|instalaci[oó]n.*hangar', desc_lower):
            return 'Operacional'
        # "Flypro hangar devolucion" → Hangar
        return 'Hangar'

    # ── 3. SEGURO ──
    if re.search(r'seguro|endoso|cuota seguro|pago seguro', desc_lower):
        # "Seguro de viaje" for Ceres → Sin clasificar
        if re.search(r'ceres', desc_lower):
            return 'Sin clasificar'
        # "Egreso Compra divisas - Pago seguro avión" → Seguro
        return 'Seguro'

    # ── 4. IMPUESTO ──
    if re.search(r'^toa\b|patente municipal|patente comercial|impuesto|iva f29|cobro iluminaci[oó]n|pago patente', desc_lower):
        return 'Impuesto'
    if re.search(r'diferencia de impuesto al lujo', desc_lower):
        return 'Impuesto'
    if re.search(r'pago contado.*operaci[oó]n renta', desc_lower):
        return 'Impuesto'

    # ── 5. INVERSIÓN (Fintual, DAP, Fondo Mutuo, Deposito a plazo, Cuenta Ahorro) ──
    if re.search(r'fintual|^dap$|cobro dap|toma dap|rescate dap|fondo mutuo|deposito a plazo|dep[oó]sito a plazo|cuenta ahorro|inversi[oó]n dep[oó]sito', desc_lower):
        return 'Inversión'
    # "Provisión Overhaul Fintual" → Inversión (it's a Fintual transfer)
    if re.search(r'provisi[oó]n overhaul fintual|overhaul.*fintual|fintual.*overhaul', desc_lower):
        return 'Inversión'
    # "Ahorro Fintual" → Inversión
    if re.search(r'ahorro fintual', desc_lower):
        return 'Inversión'
    # "Reserva Avión Fintual" or "Reserva Fintual" → Inversión
    if re.search(r'reserva.*fintual|fintual', desc_lower):
        return 'Inversión'

    # ── 6. OVERHAUL ──
    if re.search(r'overhaul|stc overhaul', desc_lower):
        return 'Overhaul'
    # Aeromundo Cambio Motor → Overhaul
    if re.search(r'cambio motor', desc_lower):
        return 'Overhaul'
    # Avinel revisión alternador → Overhaul? No, it's more Mantenimiento
    # "Aerovimath" → Overhaul (cotización 33)
    if re.search(r'aerovimath', desc_lower):
        return 'Overhaul'

    # ── 7. COMBUSTIBLE ──
    if re.search(r'combustible|avgas|litros?\b.*aqi|litros?\b.*cc-?aqi|litros?\b.*ccaqi|litros?\b.*scc[a-z]|litros?\b.*sc[a-z]{2}|club\s*aereo?\s*curacav[ií]|caiquen|esmax|^jt\b|bidón|bidon|recarga bidones', desc_lower):
        return 'Combustible'
    # Patterns like "81 litros", "93 litros", "litros avgas"
    if re.search(r'\d+[\.,]?\d*\s*litros?\b', desc_lower):
        return 'Combustible'
    # SCCV, SCVL, SCTB, SCSE, SCTL, SCIE, SCLI, SCVI fuel entries (airport codes)
    if re.search(r'^scc[vl]|^sc[a-z]{2}\s*-?\s*\d', desc_lower):
        return 'Combustible'
    # "Reembolso Fuel", "Flypro Combustible"
    if re.search(r'reembolso fuel|flypro combustible', desc_lower):
        return 'Combustible'

    # ── 8. MANTENIMIENTO (service/labor) ──
    if re.search(r'rev\.?\s*\d+\s*h|revisi[oó]n\s*\d+\s*h|cambio aceite|lavado|aspirado|mantto|mec[aá]nico|ivan oliva|iván oliva|oliva\s*(ptt|arreglo|trabajo)', desc_lower):
        return 'Mantenimiento'
    if re.search(r'aeromundo\s*rev|coordinador\s*viraje|coordinador\s*de\s*viraje|sellado|arreglo shimmy|arreglo trim|arreglo asiento|prueba de equipo|balatas', desc_lower):
        return 'Mantenimiento'
    if re.search(r'arreglo motor partida', desc_lower):
        return 'Mantenimiento'
    # "AIRC" → Mantenimiento (avionics work/labor)
    if re.search(r'^airc\b|airc\s*-?\s*(trabajo|prueba|arreglo|instalaci)', desc_lower):
        return 'Mantenimiento'
    # "Lavado AQI", "Lavado Avión"
    if re.search(r'lavado', desc_lower):
        return 'Mantenimiento'
    # "Limpieza mensual Tapia" → Mantenimiento
    if re.search(r'limpieza\s*mensual|hector\s*tapia', desc_lower):
        return 'Mantenimiento'
    # "Trabajos CC-AQI", "Trabajos Batería"
    if re.search(r'trabajos?\s*(cc-?aqi|bater[ií]a|ivan|iván|cambio|oliva)', desc_lower):
        return 'Mantenimiento'
    # "Revisión 50 Hrs", "Revisión 500 horas"
    if re.search(r'revisi[oó]n\s*\d+', desc_lower):
        return 'Mantenimiento'
    # "Mantenimiento" in description
    if re.search(r'mantenimiento', desc_lower):
        return 'Mantenimiento'
    # "Instalación" related to aircraft work
    if re.search(r'instalaci[oó]n.*antena', desc_lower):
        return 'Mantenimiento'
    # "Avinel" → Mantenimiento (alternador revision etc)
    if re.search(r'avinel', desc_lower):
        return 'Mantenimiento'
    # "Aeromundo" standalone (certificacion, payments) → Mantenimiento
    if re.search(r'^aeromundo', desc_lower):
        return 'Mantenimiento'
    # "Pedro Gotelli" → Mantenimiento (trabajos AQI)
    if re.search(r'pedro gotelli|gotelli', desc_lower):
        return 'Mantenimiento'
    # "Vuelo traslado AQI" → Mantenimiento (ferry flight for maintenance)
    if re.search(r'vuelo traslado|traslado aqi', desc_lower):
        return 'Mantenimiento'
    # "Jorge Leiva" standalone (mechanic) → Mantenimiento
    if re.search(r'^jorge\s*leiva$', desc_lower) and egreso and float(egreso) > 0:
        return 'Mantenimiento'

    # ── 9. REPUESTOS (parts/equipment purchases) ──
    if re.search(r'lideravia|aeroneed|aircraft spruce|repuesto|neum[aá]tico|bater[ií]a (airc|elt|gps)|filtro de (aceite|aire)|filtro aceite|buj[ií]a', desc_lower):
        return 'Repuestos'
    if re.search(r'hsi\b|g5\b|focos?\s*led|aud[ií]fonos|tap[ií]z|extintor|jpi\s*edm|gma\s*340|caja de audio', desc_lower):
        return 'Repuestos'
    if re.search(r'aceite\s*(tkoff|takeoff|aeroshell)|botellas?\s*aceite|w100|takeoff|tkoff', desc_lower):
        return 'Repuestos'
    if re.search(r'ducto\s*scat|oil cooler|pernos?\s*c[aá]mara|montante|manguera|empaquetadura|muffler|escape.*silenciador|silenciador', desc_lower):
        return 'Repuestos'
    if re.search(r'shimmy damper|torque link|mcfarlane|felpa rodamiento|strobe|starter|broches capota', desc_lower):
        return 'Repuestos'
    if re.search(r'funda cobertora|pasta de pulir|lineas fuel primer', desc_lower):
        return 'Repuestos'
    if re.search(r'discos?\s*de\s*freno|aceite\s*\+\s*filtro|caja\s*de\s*aceite', desc_lower):
        return 'Repuestos'
    # "Jorge Leiva" parts purchases (Fuente Poder, compra)
    if re.search(r'jorge leiva.*fuente|j\.\s*leiva\s*compra|jorge leiva.*compra|j\.\s*leiva.*compra', desc_lower):
        return 'Repuestos'
    # "Antenas Estáticas - Jorge Leiva" → Repuestos (parts)
    if re.search(r'antenas?\s*est[aá]ticas', desc_lower):
        return 'Repuestos'
    # "Filtro" standalone
    if re.search(r'^filtro\b', desc_lower):
        return 'Repuestos'
    # "6 unidades de Aeroshell"
    if re.search(r'aeroshell|w100 plus', desc_lower):
        return 'Repuestos'
    # "Aviation Parts"
    if re.search(r'aviation parts', desc_lower):
        return 'Repuestos'
    # "Batería ELT", "Batería AIRC"
    if re.search(r'bater[ií]a\b', desc_lower):
        return 'Repuestos'

    # ── 10. CERES / AEROTRUST → Sin clasificar ──
    if re.search(r'^ceres\b|ceres$|ceres\s', desc_lower) and not re.search(r'combustible.*ceres|ceres.*combustible', desc_lower):
        return 'Sin clasificar'
    if re.search(r'aerotrust|factura\s*\d+\s*aerotrust|factura\s*aerotrust', desc_lower):
        return 'Sin clasificar'
    # Facturas for Aerotrust/Stratus numbered invoices (e.g. "Facturas 575 y 576")
    if re.search(r'factura', desc_lower) and re.search(r'\d{3}', desc_lower) and not re.search(r'aeromundo', desc_lower):
        return 'Sin clasificar'
    # "Stratus SPA" payments → Sin clasificar
    if re.search(r'stratus\s*spa|aerostratus|rendiciones stratus|pagos? pendientes stratus|préstamo.*stratus|transferencia.*stratus|stratus.*devoluci', desc_lower):
        return 'Sin clasificar'
    # Mulet hours for Ceres
    if re.search(r'(j\.?\s*)?mulet|horas?\s*ceres|hrs\.?\s*ceres|horas?\s*vuelo|hrs\.?\s*vuelo', desc_lower):
        # Mulet / Ceres pilot hours → Sin clasificar
        return 'Sin clasificar'
    # Hotel/Uber for Ceres trips
    if re.search(r'hotel\s*(talca|osorno|insigne|angol|casino|chill[aá]n)|uber\s*talca|comida\s*talca|airbnb', desc_lower):
        return 'Sin clasificar'
    # Bencina y Peajes (Ceres ground transport)
    if re.search(r'bencina y peajes|bencina.*peaje', desc_lower):
        return 'Sin clasificar'
    # "Trabajo AOC Aerostratus" → Sin clasificar
    if re.search(r'aoc\s*aerostratus|trabajo.*aoc', desc_lower):
        return 'Sin clasificar'
    # "Pago inicial revisión documentos AQI" → Sin clasificar (Ceres admin)
    # Herramientas Easy for Ceres
    if re.search(r'herramientas easy', desc_lower):
        return 'Sin clasificar'

    # ── 11. PAGO PILOTO ──
    # Has a valid pilot code in cliente AND has ingreso > 0
    # EXCEPTION: Pilot codes with EGRESOS = fuel reimbursements → Combustible
    # (Before credit system, pilots paid fuel and got reimbursed)
    fuel_reimbursement_codes = {'RM', 'SV', 'PA', 'CP'}
    if cliente_clean in fuel_reimbursement_codes and egreso and float(egreso) > 0 and not (ingreso and float(ingreso) > 0):
        return 'Combustible'
    # Cer (Ceres) client code → Sin clasificar
    if cliente_clean == 'Cer':
        return 'Sin clasificar'
    # C. Piraino standalone egreso without code → Combustible (fuel reimbursement)
    if re.search(r'^c\.\s*piraino$', desc_lower) and egreso and float(egreso) > 0 and not cliente_clean:
        return 'Combustible'
    
    if cliente_clean in PILOT_CODES and ingreso and float(ingreso) > 0:
        return 'Pago piloto'
    # Negative SV ingreso entries like "S. Varas,,-1000000" also Pago piloto
    if cliente_clean == 'SV' and ingreso:
        return 'Pago piloto'
    # FlyPro entries with Fpro code
    if cliente_clean == 'Fpro':
        return 'Pago piloto'
    # USD code means Ceres payment → Sin clasificar
    if cliente_clean == 'USD':
        return 'Sin clasificar'

    # ── 12. OPERACIONAL ──
    if re.search(r'contador|s\s*y\s*p\s*(consultores|auditores)|syp|pago contador', desc_lower):
        return 'Operacional'
    if re.search(r'notarial|multa|abogad|certificados?\s*digitales', desc_lower):
        return 'Operacional'
    if re.search(r'flypro|fly\s*pro|sueldo\s*fly|horas?\s*fly\s*pro|reembolso\s*fly', desc_lower):
        return 'Operacional'
    if re.search(r'chaqueta\s*flypro|chaqueta\s*fly', desc_lower):
        return 'Operacional'
    if re.search(r'50%\s*foreflight|foreflight', desc_lower):
        return 'Operacional'
    if re.search(r'transferencia\s*(mam[aá]|pap[aá])', desc_lower):
        return 'Operacional'
    if re.search(r'reembolso\s*restaurant|pr[eé]stamo\s*pap[aá]s', desc_lower):
        return 'Operacional'
    if re.search(r'dgac|nota de d[eé]bito', desc_lower):
        return 'Operacional'
    if re.search(r'traspaso.*cta|devolucion.*cta|traspaso internet', desc_lower):
        return 'Operacional'
    if re.search(r'gabriel.*romero|trabajos gabriel|movimiento.*aviones|limpieza hangar|go detailing', desc_lower):
        return 'Operacional'
    if re.search(r'creaci[oó]n de cuenta|pago inicial cotizaci[oó]n|compresor', desc_lower):
        return 'Operacional'
    if re.search(r'^error$|^error,', desc_lower):
        return 'Operacional'
    if re.search(r'subida nieve|aseo miriam|frigobar|microondas|starlink|escalera|aspiradora|robot|sodimac|cortadora|ventilador|palmeta|enchufe|riego autom[aá]tico|instalaci[oó]n agua', desc_lower):
        return 'Operacional'
    if re.search(r'transporte$|^transf\.$|^transferencia$', desc_lower):
        return 'Operacional'
    # "Pago Erroneo" → Operacional
    if re.search(r'pago erroneo|pago err[oó]neo', desc_lower):
        return 'Operacional'
    # "Devolución - Pago Tarjeta" → Operacional
    if re.search(r'pago tarjeta credito|devoluci[oó]n.*pago tarjeta', desc_lower):
        return 'Operacional'
    # "J. Mulet - Piloto de seguridad" → Operacional
    if re.search(r'piloto de seguridad', desc_lower):
        return 'Operacional'
    # "P. del valle instrucción a Llorente" → Operacional
    if re.search(r'instrucci[oó]n\s*a\s', desc_lower):
        return 'Operacional'
    # "Trabajos" generic (Gabriel, riego, etc.)
    if re.search(r'^trabajos?\s*gabriel', desc_lower):
        return 'Operacional'
    # "16,8 litros AQI - Reembolso J. Pizarro" → Combustible actually
    # Covered above by litros pattern

    # ── 13. PAGO PILOTO — names without code in cliente column ──
    # Older entries have pilot names in description but no code
    old_pilot_names = [
        r'^r\.\s*alvarez', r'^a\.\s*fernandez', r'^amir\s*kia', r'^a\.\s*kia',
        r'^r\.\s*castro', r'^f\.\s*hidalgo', r'^g\.\s*latorre', r'^a\.\s*torrealba',
        r'^d\.\s*ross', r'^f\.\s*hernandez', r'^v\.\s*asenjo', r'^j\.\s*chiang',
        r'^p\.\s*agliati', r'^j\.\s*mathew', r'^v\.\s*amengual', r'^n\.\s*elias',
        r'^n\.\s*vega', r'^l\.\s*montoya', r'^jesus\s*bermedo',
        r'^c\.\s*pomes', r'^d\.\s*gutierrez$', r'^j\.\s*varas$',
        r'^r\.\s*fuentes$', r'^n\.\s*espinoza$', r'^v\.\s*ortiz$',
        r'^c\.\s*piraino$',
        r'^pablo\s*jimenez', r'^alfredo\s*saavedra', r'^daniel\s*osorio',
        r'^daniel\s*roco', r'^r\.\s*mejia$', r'^s\.\s*varas$',
    ]
    for pat in old_pilot_names:
        if re.search(pat, desc_lower) and ingreso and float(ingreso) > 0:
            return 'Pago piloto'

    # ── 14. SPECIFIC DESCRIPTIONS ──
    # "FlyPro" ingresos without code → Operacional
    if re.search(r'^flypro$|^flypro\s', desc_lower):
        return 'Operacional'

    # "Pago Contador S&P" → Operacional
    if re.search(r'pago contador', desc_lower):
        return 'Operacional'

    # "Transferencia a Stratus SPA" → Sin clasificar
    if re.search(r'stratus', desc_lower):
        return 'Sin clasificar'

    # "Instrucción S. Varas" (egreso for safety pilot) → Operacional
    if re.search(r'instrucci[oó]n\s*s\.?\s*varas|s\.\s*varas\s*instrucci[oó]n|instrucci[oó]n.*s\.\s*varas|pago\s*s\.\s*varas\s*vuelo', desc_lower):
        return 'Operacional'
    # "Instruccion I. Opazo  S. Varas" → Operacional
    if re.search(r'instrucci[oó]n', desc_lower):
        return 'Operacional'

    # "Rescate Fondos Mutuos" → Inversión
    if re.search(r'rescate\s*fondos?\s*mutuo|fondos?\s*mutuo', desc_lower):
        return 'Inversión'

    # "Colmena" → Operacional (AFP)
    if re.search(r'colmena', desc_lower):
        return 'Operacional'

    # "Reembolso" generic (Neumático, etc.)
    if re.search(r'reembolso\s*neum', desc_lower):
        return 'Repuestos'
    if re.search(r'reembolso', desc_lower):
        return 'Operacional'

    # Fuel patterns with "lts" abbreviation
    if re.search(r'\d+\s*lts?\s*(aqi|avgas|cc)', desc_lower):
        return 'Combustible'

    # ── 15. CATCH-ALL for known patterns ──
    # Ceres-related descriptions
    if re.search(r'ceres', desc_lower):
        return 'Sin clasificar'

    # "Diferencia" (adjustment entries for Ceres)
    if re.search(r'diferencia\s*\d', desc_lower):
        return 'Sin clasificar'

    # "FlyPro" bono, horas → Operacional
    if re.search(r'bono|flypr', desc_lower):
        return 'Operacional'

    # "A. Corsi" entries → Operacional (reembolso/sueldo)
    if re.search(r'a\.\s*corsi', desc_lower):
        return 'Operacional'

    # "Prestamos Stratus" → Sin clasificar  
    if re.search(r'pr[eé]stamo', desc_lower):
        # "Prestamo JT" → Operacional
        if re.search(r'prestamo\s*jt', desc_lower):
            return 'Operacional'
        return 'Sin clasificar'

    # "Transf." → Operacional
    if re.search(r'^transf', desc_lower):
        return 'Operacional'

    # "ACPDV" → Operacional
    if re.search(r'acpdv', desc_lower):
        return 'Operacional'

    # "Depósito en Efectivo" → Operacional
    if re.search(r'dep[oó]sito en efectivo', desc_lower):
        return 'Operacional'

    # "Aporte Capital" → Operacional
    if re.search(r'aporte\s*capital', desc_lower):
        return 'Operacional'

    # "Pago a R. Fuentes por vuelo" → Operacional
    if re.search(r'pago a .* por vuelo', desc_lower):
        return 'Operacional'

    # "Claudio Guajardo" → Mantenimiento (AIRC related)
    if re.search(r'claudio guajardo', desc_lower):
        return 'Mantenimiento'

    # Travel expenses: Food, Transportation, Perdiem
    if re.search(r'^food\b|^transportation\b|perdiem|transvip', desc_lower):
        return 'Sin clasificar'

    # "Antena GPS" → Repuestos
    if re.search(r'antena gps', desc_lower):
        return 'Repuestos'

    # "Wifi upgrade" → Operacional
    if re.search(r'wifi\s*upgrade', desc_lower):
        return 'Operacional'

    # "Trabajos Plataforma" → Operacional
    if re.search(r'trabajos?\s*plataforma', desc_lower):
        return 'Operacional'

    # "Airbnb" → Sin clasificar (Ceres)
    if re.search(r'airbnb', desc_lower):
        return 'Sin clasificar'

    # "Pedro Gotelli" → Mantenimiento (trabajos AQI)
    if re.search(r'pedro gotelli|gotelli', desc_lower):
        return 'Mantenimiento'

    # "Vuelo traslado AQI" → Mantenimiento (ferry flight for maintenance)
    if re.search(r'vuelo traslado|traslado aqi', desc_lower):
        return 'Mantenimiento'

    # "Pago inicial revisión" → Operacional
    if re.search(r'pago inicial revisi[oó]n', desc_lower):
        return 'Operacional'

    # "Prestamo JT" → Operacional
    if re.search(r'prestamo\s*jt', desc_lower):
        return 'Operacional'

    # ── 16. FINAL FALLBACK ──
    return 'Sin clasificar'


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Find header row and column mapping
    # Columns: B=Correlativo, C=Fecha, D=Descripción, E=Egreso, F=Ingreso, G=Saldo, H=Tipo, I=Cliente
    # Based on earlier analysis, data starts at row 2, columns B-I
    
    header_row = 1
    tipo_col = 8   # Column H (1-indexed)
    desc_col = 4   # Column D
    egreso_col = 5 # Column E
    ingreso_col = 6 # Column F
    saldo_col = 7   # Column G
    cliente_col = 9 # Column I

    stats = {}
    changed = 0
    total = 0
    kept = 0

    for row_num in range(2, ws.max_row + 1):
        desc = ws.cell(row=row_num, column=desc_col).value
        if not desc:
            continue
        
        total += 1
        desc = str(desc).strip()
        egreso = ws.cell(row=row_num, column=egreso_col).value
        ingreso = ws.cell(row=row_num, column=ingreso_col).value
        tipo_existing = ws.cell(row=row_num, column=tipo_col).value
        cliente = ws.cell(row=row_num, column=cliente_col).value

        tipo_existing_str = str(tipo_existing).strip() if tipo_existing else ''
        cliente_str = str(cliente).strip() if cliente else ''

        new_tipo = classify_row(desc, egreso, ingreso, tipo_existing_str, cliente_str)

        if tipo_existing_str and tipo_existing_str == new_tipo:
            kept += 1
        elif tipo_existing_str != new_tipo:
            changed += 1

        ws.cell(row=row_num, column=tipo_col).value = new_tipo
        stats[new_tipo] = stats.get(new_tipo, 0) + 1

    wb.save(XLSX_PATH)

    print(f"\n{'='*50}")
    print(f"Classification complete!")
    print(f"{'='*50}")
    print(f"Total rows processed: {total}")
    print(f"Kept existing tipo:   {kept}")
    print(f"Changed/assigned:     {changed}")
    print(f"\nBreakdown by tipo:")
    print(f"{'-'*40}")
    for tipo, count in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"  {tipo:<20} {count:>5}")
    print(f"{'-'*40}")
    print(f"  {'TOTAL':<20} {sum(stats.values()):>5}")


if __name__ == '__main__':
    main()
